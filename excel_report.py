# ============================================================
# EXCEL REPORT GENERATOR
# Exports a color-coded .xlsx report using openpyxl
# Run: python3 excel_report.py
# ============================================================

import openpyxl
from openpyxl.styles import (
    PatternFill,   # background cell color
    Font,          # bold, size, color of text
    Alignment,     # center / left align text
    Border,        # cell borders
    Side           # border line style
)
from openpyxl.utils import get_column_letter  # converts 1 → "A", 2 → "B" etc.
from openpyxl.chart import BarChart, Reference  # for the embedded chart
import pandas as pd
import datetime


# ============================================================
# CONFIGURATION
# ============================================================

DATA_PATH         = "data/students.csv"
OUTPUT_PATH       = "student_report.xlsx"
TOP_AVG           = 80
TOP_ATT           = 85
RISK_AVG          = 50
RISK_ATT          = 65


# ============================================================
# COLORS  (Excel uses ARGB hex — "FF" prefix = fully opaque)
# ============================================================

COLOR_GREEN_BG    = "FFE1F5EE"   # light green fill — top performers
COLOR_GREEN_FONT  = "FF0F6E56"   # dark green text
COLOR_RED_BG      = "FFFCEBEB"   # light red fill — at-risk
COLOR_RED_FONT    = "FFA32D2D"   # dark red text
COLOR_BLUE_BG     = "FFE6F1FB"   # light blue fill — average
COLOR_BLUE_FONT   = "FF185FA5"   # dark blue text
COLOR_HEADER_BG   = "FF1D9E75"   # green header row background
COLOR_HEADER_FONT = "FFFFFFFF"   # white header text
COLOR_TITLE_FONT  = "FF0F6E56"   # dark green for sheet title
COLOR_GRAY_BG     = "FFF5F5F3"   # alternating row background


# ============================================================
# HELPER: make a fill object from a hex color
# ============================================================

def make_fill(hex_color):
    return PatternFill(start_color=hex_color,
                       end_color=hex_color,
                       fill_type="solid")


# ============================================================
# HELPER: make a thin border for all 4 sides of a cell
# ============================================================

def make_border():
    thin = Side(style="thin", color="FFD3D1C7")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================
# HELPER: apply styles to a single cell
# ============================================================

def style_cell(cell, bold=False, font_color="FF1A1A18",
               fill=None, align="left", size=11):
    cell.font      = Font(bold=bold, color=font_color, size=size,
                          name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = make_border()
    if fill:
        cell.fill  = fill


# ============================================================
# LOAD AND PREPARE DATA
# ============================================================

df = pd.read_csv(DATA_PATH)
df["Average"]    = df[["Math", "Science", "English"]].mean(axis=1).round(2)
df["Attendance"] = df["Attendance"].astype(float)

# Assign status to every student
def get_status(row):
    if row["Average"] >= TOP_AVG and row["Attendance"] >= TOP_ATT:
        return "Top Performer"
    elif row["Average"] < RISK_AVG or row["Attendance"] < RISK_ATT:
        return "At-Risk"
    else:
        return "Average"

df["Status"] = df.apply(get_status, axis=1)
df_sorted    = df.sort_values("Average", ascending=False).reset_index(drop=True)


# ============================================================
# CREATE WORKBOOK WITH 3 SHEETS
# ============================================================

wb = openpyxl.Workbook()

# openpyxl creates one default sheet — rename it
wb.active.title = "Student Report"
ws1 = wb.active
ws2 = wb.create_sheet("Summary")
ws3 = wb.create_sheet("At-Risk Students")


# ============================================================
# SHEET 1: FULL STUDENT REPORT
# ============================================================

# ── Title row ─────────────────────────────────────────────
ws1.merge_cells("A1:H1")  # merge across all 8 columns
title_cell = ws1["A1"]
title_cell.value     = "Student Performance Analyzer — Full Report"
title_cell.font      = Font(bold=True, size=16, color=COLOR_TITLE_FONT, name="Calibri")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
title_cell.fill      = make_fill("FFF0FDF4")
ws1.row_dimensions[1].height = 36

# ── Subtitle row (generated date) ─────────────────────────
ws1.merge_cells("A2:H2")
now = datetime.datetime.now().strftime("%d %B %Y, %I:%M %p")
sub_cell = ws1["A2"]
sub_cell.value     = f"Generated on: {now}"
sub_cell.font      = Font(size=10, color="FF888780", name="Calibri")
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 20

# ── Empty spacer row ───────────────────────────────────────
ws1.row_dimensions[3].height = 8

# ── Header row ────────────────────────────────────────────
headers = ["#", "Name", "Math", "Science", "English",
           "Average", "Attendance %", "Status"]
header_fill = make_fill(COLOR_HEADER_BG)

for col_num, header in enumerate(headers, start=1):
    cell = ws1.cell(row=4, column=col_num, value=header)
    style_cell(cell,
               bold=True,
               font_color=COLOR_HEADER_FONT,
               fill=header_fill,
               align="center",
               size=11)

ws1.row_dimensions[4].height = 24

# ── Data rows ─────────────────────────────────────────────
for row_num, (_, student) in enumerate(df_sorted.iterrows(), start=5):

    status = student["Status"]

    # Pick colors based on status
    if status == "Top Performer":
        bg_fill   = make_fill(COLOR_GREEN_BG)
        font_color = COLOR_GREEN_FONT
    elif status == "At-Risk":
        bg_fill   = make_fill(COLOR_RED_BG)
        font_color = COLOR_RED_FONT
    else:
        # Alternate gray/white for average students — easier to read
        if row_num % 2 == 0:
            bg_fill = make_fill(COLOR_GRAY_BG)
        else:
            bg_fill = make_fill("FFFFFFFF")
        font_color = "FF1A1A18"

    row_data = [
        row_num - 4,                    # rank number
        student["Name"],
        int(student["Math"]),
        int(student["Science"]),
        int(student["English"]),
        student["Average"],
        student["Attendance"],
        status
    ]

    for col_num, value in enumerate(row_data, start=1):
        cell = ws1.cell(row=row_num, column=col_num, value=value)
        align = "center" if col_num != 2 else "left"
        style_cell(cell,
                   font_color=font_color,
                   fill=bg_fill,
                   align=align)

    ws1.row_dimensions[row_num].height = 20

# ── Column widths ─────────────────────────────────────────
col_widths = [5, 22, 8, 10, 10, 10, 14, 16]
for i, width in enumerate(col_widths, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = width

# ── Freeze top 4 rows so header stays visible when scrolling ──
ws1.freeze_panes = "A5"


# ============================================================
# SHEET 2: SUMMARY STATISTICS
# ============================================================

summary_title = ws2["A1"]
ws2.merge_cells("A1:C1")
summary_title.value     = "Class Summary"
summary_title.font      = Font(bold=True, size=14, color=COLOR_TITLE_FONT, name="Calibri")
summary_title.alignment = Alignment(horizontal="center", vertical="center")
summary_title.fill      = make_fill("FFF0FDF4")
ws2.row_dimensions[1].height = 30

# Summary data rows
subject_avgs = {
    "Math"    : df["Math"].mean().round(2),
    "Science" : df["Science"].mean().round(2),
    "English" : df["English"].mean().round(2),
}

pass_rate   = round(len(df[df["Average"] >= RISK_AVG]) / len(df) * 100, 1)
att_rate    = round(len(df[df["Attendance"] >= 65]) / len(df) * 100, 1)
top_count   = len(df[df["Status"] == "Top Performer"])
risk_count  = len(df[df["Status"] == "At-Risk"])
correlation = df["Attendance"].corr(df["Average"]).round(2)

summary_rows = [
    ("Metric", "Value", ""),                          # sub-header
    ("Total students",         len(df),               ""),
    ("Class average",          df["Average"].mean().round(2), "/ 100"),
    ("Pass rate",              f"{pass_rate}%",        "avg >= 50"),
    ("Good attendance rate",   f"{att_rate}%",         "attendance >= 65%"),
    ("Top performers",         top_count,              "avg >= 80 & attend >= 85%"),
    ("At-risk students",       risk_count,             "avg < 50 OR attend < 65%"),
    ("Math average",           subject_avgs["Math"],   "/ 100"),
    ("Science average",        subject_avgs["Science"],"/ 100"),
    ("English average",        subject_avgs["English"],"/ 100"),
    ("Attendance correlation", correlation,            "with average marks"),
]

for row_num, (metric, value, note) in enumerate(summary_rows, start=2):
    is_subheader = metric == "Metric"

    for col, val in zip([1, 2, 3], [metric, value, note]):
        cell = ws2.cell(row=row_num, column=col, value=val)

        if is_subheader:
            style_cell(cell, bold=True,
                       font_color=COLOR_HEADER_FONT,
                       fill=make_fill(COLOR_HEADER_BG),
                       align="center")
        else:
            fill = make_fill(COLOR_GRAY_BG) if row_num % 2 == 0 else make_fill("FFFFFFFF")
            style_cell(cell, fill=fill, align="left" if col == 1 else "center")

    ws2.row_dimensions[row_num].height = 20

ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 28

# ── Embed a bar chart of subject averages into Sheet 2 ────
chart      = BarChart()
chart.type = "col"                        # vertical bars
chart.title       = "Subject Averages"
chart.y_axis.title = "Average Marks"
chart.x_axis.title = "Subject"
chart.style       = 10
chart.width       = 15
chart.height      = 12

# Data reference: rows 9–11 (Math, Science, English averages), column B
data = Reference(ws2, min_col=2, min_row=9, max_row=11)
cats = Reference(ws2, min_col=1, min_row=9, max_row=11)
chart.add_data(data)
chart.set_categories(cats)
chart.shape = 4

ws2.add_chart(chart, "E2")  # place chart starting at cell E2


# ============================================================
# SHEET 3: AT-RISK STUDENTS (focused view)
# ============================================================

ws3.merge_cells("A1:F1")
at_risk_title = ws3["A1"]
at_risk_title.value     = "At-Risk Students — Requires Immediate Attention"
at_risk_title.font      = Font(bold=True, size=13, color=COLOR_RED_FONT, name="Calibri")
at_risk_title.alignment = Alignment(horizontal="center", vertical="center")
at_risk_title.fill      = make_fill("FFFCEBEB")
ws3.row_dimensions[1].height = 30

# Header
at_risk_headers = ["Name", "Math", "Science", "English", "Average", "Attendance %"]
for col_num, h in enumerate(at_risk_headers, start=1):
    cell = ws3.cell(row=2, column=col_num, value=h)
    style_cell(cell, bold=True,
               font_color=COLOR_HEADER_FONT,
               fill=make_fill("FFE24B4A"),
               align="center")

# At-risk student rows
at_risk_df = df[df["Status"] == "At-Risk"].sort_values("Average")
for row_num, (_, student) in enumerate(at_risk_df.iterrows(), start=3):
    row_data = [
        student["Name"],
        int(student["Math"]),
        int(student["Science"]),
        int(student["English"]),
        student["Average"],
        student["Attendance"]
    ]
    for col_num, value in enumerate(row_data, start=1):
        cell = ws3.cell(row=row_num, column=col_num, value=value)
        style_cell(cell,
                   fill=make_fill(COLOR_RED_BG),
                   font_color=COLOR_RED_FONT,
                   align="left" if col_num == 1 else "center")
    ws3.row_dimensions[row_num].height = 20

# Column widths
for col, width in zip([1,2,3,4,5,6], [22,8,10,10,10,14]):
    ws3.column_dimensions[get_column_letter(col)].width = width


# ============================================================
# SAVE THE WORKBOOK
# ============================================================

wb.save(OUTPUT_PATH)
print(f"Excel report saved: {OUTPUT_PATH}")
print(f"  Sheet 1 — Full student report ({len(df)} students, color-coded)")
print(f"  Sheet 2 — Summary statistics + bar chart")
print(f"  Sheet 3 — At-risk students ({risk_count} flagged)")